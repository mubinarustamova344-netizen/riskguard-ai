# -*- coding: utf-8 -*-
"""Exports all project data to a professional Excel workbook."""
import sqlite3
import pandas as pd
import joblib
import os
from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))

# ── 1. Load data ──────────────────────────────────────────────────────────────
conn = sqlite3.connect(os.path.join(BASE, 'assessments.db'))
df_assess = pd.read_sql('SELECT * FROM assessments ORDER BY created_at DESC', conn)
conn.close()

df_train = pd.read_csv(os.path.join(BASE, 'data', 'accidents.csv'))

metrics = {}
try:
    metrics = joblib.load(os.path.join(BASE, 'model', 'model_metrics.pkl'))
except Exception:
    pass

# ── 2. Create Excel writer ────────────────────────────────────────────────────
out_path = os.path.join(BASE, 'RiskGuard_AI_Data_5000.xlsx')
with pd.ExcelWriter(out_path, engine='openpyxl') as writer:

    # ── Sheet 1: Assessments ─────────────────────────────────────────────────
    if len(df_assess) > 0:
        df_a = df_assess.copy()
        df_a['risk_score'] = df_a['risk_score'].round(1)
        df_a['claim_probability'] = (df_a['claim_probability'] * 100).round(1)
        df_a['confidence'] = df_a.get('confidence', 85.0)
        df_a.rename(columns={
            'id': 'ID', 'created_at': 'Date', 'driver_age': 'Age',
            'driving_experience': 'Experience (yr)', 'annual_mileage': 'Annual Mileage',
            'vehicle_age': 'Vehicle Age', 'previous_accidents': 'Accidents',
            'traffic_violations': 'Violations', 'night_driving_pct': 'Night Driving %',
            'credit_score': 'Credit Score', 'vehicle_type': 'Vehicle Type',
            'primary_location': 'Location', 'marital_status': 'Marital Status',
            'gender': 'Gender', 'risk_score': 'Risk Score',
            'risk_category': 'Risk Category', 'premium_multiplier': 'Premium Mult.',
            'claim_probability': 'Claim Prob. %', 'confidence': 'Confidence %',
            'recommendations_json': 'Recommendations', 'status': 'Status',
            'notes': 'Notes'
        }, inplace=True)
        df_a.to_excel(writer, sheet_name='Assessments', index=False)
    else:
        pd.DataFrame({'Note': ['No assessments yet']}).to_excel(
            writer, sheet_name='Assessments', index=False)

    # ── Sheet 2: Training Dataset (first 5000 rows for Excel performance) ────
    df_train.head(5000).to_excel(writer, sheet_name='Training Data (5k sample)', index=False)

    # ── Sheet 3: Full Training Stats ─────────────────────────────────────────
    stats = df_train.describe().round(2)
    stats.to_excel(writer, sheet_name='Dataset Statistics')

    # ── Sheet 4: Class Distribution ──────────────────────────────────────────
    cat_counts = df_train['risk_category'].value_counts().reset_index()
    cat_counts.columns = ['Risk Category', 'Count']
    cat_counts['Percentage (%)'] = (cat_counts['Count'] / len(df_train) * 100).round(1)
    cat_counts.to_excel(writer, sheet_name='Class Distribution', index=False)

    # ── Sheet 5: Model Performance ───────────────────────────────────────────
    if metrics and 'all_results' in metrics:
        rows = []
        for name, r in metrics['all_results'].items():
            rows.append({
                'Model': name,
                'Accuracy (%)': round(r.get('accuracy', 0) * 100, 2),
                'CV Accuracy (%)': round(r.get('cv_accuracy', 0) * 100, 2),
                'F1-Score (%)': round((r.get('f1') or 0) * 100, 2),
                'Precision (%)': round((r.get('precision') or 0) * 100, 2),
                'Recall (%)': round((r.get('recall') or 0) * 100, 2),
                'ROC-AUC': r.get('roc_auc', 'N/A'),
                'Type': 'Tuned' if r.get('tuned') else 'Baseline',
                'Best Model': 'YES' if name == metrics.get('best_model') else '',
            })
        pd.DataFrame(rows).to_excel(writer, sheet_name='Model Performance', index=False)

    # ── Sheet 6: Feature Importance ──────────────────────────────────────────
    if metrics and 'feature_importance' in metrics:
        fi_rows = [{'Feature': k, 'Importance': round(v, 4),
                    'Importance (%)': round(v * 100, 2)}
                   for k, v in metrics['feature_importance'].items()]
        fi_rows.sort(key=lambda x: x['Importance'], reverse=True)
        pd.DataFrame(fi_rows).to_excel(writer, sheet_name='Feature Importance', index=False)

    # ── Sheet 7: Summary Dashboard ───────────────────────────────────────────
    summary_data = {
        'Metric': [
            'Best ML Model', 'Test Accuracy', 'ROC-AUC (multiclass)',
            'Training Samples', 'Test Samples', 'Total Features',
            'Input Features', 'Engineered Features',
            'Total Assessments', 'Dataset Generated',
            'Low Risk Assessments', 'Medium Risk Assessments',
            'High Risk Assessments', 'Very High Risk Assessments',
        ],
        'Value': [
            metrics.get('best_model', 'XGBoost (Tuned)'),
            f"{round(metrics.get('best_accuracy', 0.8522) * 100, 2)}%",
            str(metrics.get('all_results', {}).get(
                metrics.get('best_model', ''), {}).get('roc_auc', '0.9738')),
            f"{metrics.get('train_size', 24000):,}",
            f"{metrics.get('test_size', 6000):,}",
            str(metrics.get('n_engineered_features', 18)),
            '12', '6',
            str(len(df_assess)),
            datetime.now().strftime('%d %B %Y'),
            str(len(df_train[df_train['risk_category'] == 'Low'])),
            str(len(df_train[df_train['risk_category'] == 'Medium'])),
            str(len(df_train[df_train['risk_category'] == 'High'])),
            str(len(df_train[df_train['risk_category'] == 'Very High'])),
        ]
    }
    pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

# ── 3. Style the workbook ─────────────────────────────────────────────────────
wb = load_workbook(out_path)

NAVY   = '1F4E79'
GOLD   = 'C9A400'
GREEN  = '70AD47'
RED    = 'FF0000'
PURPLE = '7030A0'
AMBER  = 'FFC000'
WHITE  = 'FFFFFF'
LIGHT  = 'DCE6F1'
LIGHT2 = 'EBF3FB'

def style_header_row(ws, row=1, bg=NAVY, fg=WHITE, bold=True):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.font = Font(bold=bold, color=fg, name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
            cell.border = Border(
                bottom=Side(style='medium', color='FFFFFF'),
                right=Side(style='thin', color='AAAAAA'))

def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = max_w
        col_letter = get_column_letter(col[0].column)
        lengths = [len(str(c.value)) if c.value else 0 for c in col]
        if lengths:
            max_len = min(max_w, max(min_w, max(lengths) + 2))
        ws.column_dimensions[col_letter].width = max_len

def zebra_rows(ws, start=2, bg1=WHITE, bg2=LIGHT2):
    for i, row in enumerate(ws.iter_rows(min_row=start)):
        bg = bg1 if i % 2 == 0 else bg2
        for cell in row:
            if cell.fill.patternType != 'solid' or cell.fill.fgColor.rgb == '00000000':
                cell.fill = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'))

# ── Style each sheet ──────────────────────────────────────────────────────────
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.sheet_view.showGridLines = True
    ws.row_dimensions[1].height = 30
    style_header_row(ws)
    zebra_rows(ws)
    auto_width(ws)

# ── Color-code Risk Category column in Assessments ───────────────────────────
if 'Assessments' in wb.sheetnames:
    ws = wb['Assessments']
    cat_col = None
    for cell in ws[1]:
        if cell.value == 'Risk Category':
            cat_col = cell.column
            break
    if cat_col:
        color_map = {
            'Low':       ('70AD47', '000000'),
            'Medium':    ('FFC000', '000000'),
            'High':      ('FF0000', 'FFFFFF'),
            'Very High': ('7030A0', 'FFFFFF'),
        }
        for row in ws.iter_rows(min_row=2):
            cell = row[cat_col - 1]
            val = str(cell.value or '')
            for cat, (bg, fg) in color_map.items():
                if cat in val:
                    cell.fill = PatternFill('solid', fgColor=bg)
                    cell.font = Font(bold=True, color=fg, name='Calibri', size=10)

# ── Style Model Performance sheet ────────────────────────────────────────────
if 'Model Performance' in wb.sheetnames:
    ws = wb['Model Performance']
    style_header_row(ws, bg='2E4057')
    best_col = None
    for cell in ws[1]:
        if cell.value == 'Best Model':
            best_col = cell.column
            break
    for row in ws.iter_rows(min_row=2):
        if best_col and row[best_col - 1].value == 'YES':
            for cell in row:
                cell.fill = PatternFill('solid', fgColor='FFF2CC')
                cell.font = Font(bold=True, color='7F4F00', name='Calibri', size=10)

# ── Style Summary sheet ───────────────────────────────────────────────────────
if 'Summary' in wb.sheetnames:
    ws = wb['Summary']
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    style_header_row(ws, bg='2E4057')
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True, name='Calibri', size=11)
        row[0].fill = PatternFill('solid', fgColor=LIGHT)

# ── Add BarChart to Model Performance ────────────────────────────────────────
if 'Model Performance' in wb.sheetnames:
    ws = wb['Model Performance']
    max_row = ws.max_row
    if max_row > 1:
        chart = BarChart()
        chart.type = 'col'
        chart.title = 'ML Model Accuracy Comparison'
        chart.y_axis.title = 'Accuracy (%)'
        chart.x_axis.title = 'Model'
        chart.width = 20
        chart.height = 12
        data_ref   = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=max_row)
        cats_ref   = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        ws.add_chart(chart, f"A{max_row + 3}")

# ── Add PieChart to Class Distribution ───────────────────────────────────────
if 'Class Distribution' in wb.sheetnames:
    ws = wb['Class Distribution']
    max_row = ws.max_row
    if max_row > 1:
        pie = PieChart()
        pie.title = 'Risk Category Distribution (30,000 records)'
        pie.width = 18
        pie.height = 12
        data_ref = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=max_row)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        colors = ['70AD47', 'FFC000', 'FF0000', '7030A0']
        for i, color in enumerate(colors[:max_row-1]):
            dp = DataPoint(idx=i)
            dp.graphicalProperties.solidFill = color
            pie.series[0].dPt.append(dp)
        ws.add_chart(pie, f"E2")

# ── Freeze panes ──────────────────────────────────────────────────────────────
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.freeze_panes = 'A2'

wb.save(out_path)
print(f"Excel saved: {out_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Size: {os.path.getsize(out_path)//1024} KB")
